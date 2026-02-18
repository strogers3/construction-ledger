import datetime
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
import openpyxl
from ledger.models import Supplier, TypeDescription, ConstructionEntry


class Command(BaseCommand):
    help = 'Import construction data from Construction 2022.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='Construction 2022.xlsx',
            help='Path to the Excel file',
        )

    def handle(self, *args, **options):
        filepath = options['file']
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['Const Actual']

        # Pre-build TypeDescription lookup from spreadsheet data
        type_map = {}  # code -> TypeDescription instance
        created_count = 0
        skipped_count = 0

        for row_num in range(9, ws.max_row + 1):
            type_code = ws.cell(row=row_num, column=21).value
            typ_desc = ws.cell(row=row_num, column=22).value

            if type_code and typ_desc:
                code_str = str(type_code).strip()
                desc_str = str(typ_desc).strip()
                if code_str in ('#VALUE!', '') or desc_str in ('#VALUE!', '0', ''):
                    continue
                if code_str not in type_map:
                    obj, _ = TypeDescription.objects.get_or_create(
                        code=code_str, defaults={'description': desc_str}
                    )
                    type_map[code_str] = obj

        self.stdout.write(f"Type descriptions loaded: {len(type_map)}")

        for row_num in range(9, ws.max_row + 1):
            date_val = ws.cell(row=row_num, column=1).value
            desc_val = ws.cell(row=row_num, column=2).value

            # Skip completely empty rows
            if date_val is None and desc_val is None:
                skipped_count += 1
                continue

            # Parse date
            if isinstance(date_val, datetime.datetime):
                date_val = date_val.date()
            elif isinstance(date_val, str):
                try:
                    date_val = datetime.datetime.strptime(date_val, '%Y-%m-%d').date()
                except ValueError:
                    date_val = None

            # Parse supplier
            supplier_name = ws.cell(row=row_num, column=5).value
            supplier = None
            if supplier_name:
                supplier_name = str(supplier_name).strip()
                if supplier_name:
                    supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

            # Parse type description
            type_code = ws.cell(row=row_num, column=21).value
            type_desc_obj = None
            if type_code:
                type_desc_obj = type_map.get(str(type_code).strip())

            # Helper for decimal fields
            def to_decimal(col):
                val = ws.cell(row=row_num, column=col).value
                if val is None:
                    return None
                try:
                    return Decimal(str(val))
                except (InvalidOperation, ValueError):
                    return None

            # Clean string helper
            def clean_str(col, max_len=200):
                val = ws.cell(row=row_num, column=col).value
                if val is None:
                    return ''
                return str(val).strip()[:max_len]

            entry = ConstructionEntry(
                date=date_val,
                description=clean_str(2, 500),
                stage=clean_str(3, 20),
                lc_stage=clean_str(4, 20),
                supplier=supplier,
                estimate=to_decimal(6),
                qty=to_decimal(7),
                supplies_cost=to_decimal(8),
                tax_fees=to_decimal(9),
                cost=to_decimal(10),
                invoiced_amt=to_decimal(11),
                posted=clean_str(12, 10),
                lm=clean_str(13, 5),
                supervisor=clean_str(14),
                invoice_number=clean_str(15, 50),
                delivery_type=clean_str(16, 20),
                materials=clean_str(17),
                book_number=clean_str(18, 20),
                notes=clean_str(19, 5000),
                type_description=type_desc_obj,
            )
            entry.save()
            created_count += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import complete: {created_count} entries created, {skipped_count} empty rows skipped."
        ))
